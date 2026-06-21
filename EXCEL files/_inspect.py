# -*- coding: utf-8 -*-
import openpyxl, glob, os, io, sys
out = io.open(r"C:\Users\num_s\KLH-ERP\EXCEL files\_inspect_out.txt", "w", encoding="utf-8")
from openpyxl.utils import get_column_letter

files = sorted(glob.glob(r"C:\Users\num_s\KLH-ERP\EXCEL files\*.xlsx"))
for f in files:
    out.write("="*80 + "\nFILE: " + os.path.basename(f) + "\n")
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    # inspect first 2 "data" sheets in detail
    shown = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
        # heuristic: a data sheet has many non-empty cells in cols
        nonempty = sum(1 for r in rows for c in r if c not in (None, ""))
        if nonempty < 8:
            continue
        out.write("\n--- SHEET: %s (dims=%s) ---\n" % (sn, ws.calculate_dimension()))
        for i, r in enumerate(rows, 1):
            cells = []
            for j, c in enumerate(r[:18]):
                if c is not None and str(c).strip() != "":
                    cells.append("%s=%s" % (get_column_letter(j+1), str(c)[:20]))
            if cells:
                out.write("r%d: %s\n" % (i, " | ".join(cells)))
        shown += 1
        if shown >= 2:
            break
    wb.close()
out.close()
print("DONE")
