# -*- coding: utf-8 -*-
import openpyxl, glob, os, re, csv

def num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    m = re.search(r'(\d+(?:\.\d+)?)', str(v).replace(',',''))
    return float(m.group(1)) if m else None

files = sorted(glob.glob(r"C:\Users\num_s\KLH-ERP\EXCEL files\*.xlsx"))
out = open(r"C:\Users\num_s\KLH-ERP\EXCEL files\_pricebook.csv","w",encoding="utf-8-sig",newline="")
w = csv.writer(out)
w.writerow(["file","sheet","name","size","pack","retail_raw","retail_num","whole_raw","whole_num","cost_case_raw","cost_unit_raw","cost_unit_num","date"])

stats = {"sheets":0,"rows":0,"named":0,"has_cost":0,"has_retail":0}
for f in files:
    fn = os.path.basename(f).split(".")[0]
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row or 1, values_only=True))
        # find header row (B == "รายการ")
        hdr = None
        for i,r in enumerate(rows[:12]):
            if len(r)>1 and r[1] and "รายการ" in str(r[1]): hdr=i; break
            if len(r)>0 and r[0] and "รายการ" in str(r[0]): hdr=i; break
        if hdr is None: continue
        # detect E/F meaning from the next row (row hdr+1): E label
        e_is_retail = True
        if hdr+1 < len(rows):
            r6 = rows[hdr+1]
            e_lbl = str(r6[4] or "") if len(r6)>4 else ""
            if "ส่ง" in e_lbl: e_is_retail=False
        stats["sheets"]+=1
        last_name=""
        for r in rows[hdr+2:]:
            if not r or len(r)<11: continue
            B = str(r[1]).strip() if r[1] else ""
            C = str(r[2]).strip() if len(r)>2 and r[2] else ""
            D = str(r[3]).strip() if len(r)>3 and r[3] else ""
            E = r[4] if len(r)>4 else None
            F = r[5] if len(r)>5 else None
            G = r[6] if len(r)>6 else None
            J = r[9] if len(r)>9 else None
            K = r[10] if len(r)>10 else None
            # name carry-down (size variant rows)
            if B: last_name=B
            name = B or (last_name if C else "")
            if not name: continue
            # skip supplier header lines (no size, no prices)
            if not C and E is None and F is None and K is None: continue
            retail = E if e_is_retail else F
            whole  = F if e_is_retail else E
            stats["rows"]+=1
            if B: stats["named"]+=1
            cu = num(K)
            rn = num(retail)
            if cu: stats["has_cost"]+=1
            if rn: stats["has_retail"]+=1
            w.writerow([fn, sn, name, C, D, retail, rn, whole, num(whole), J, K, cu,
                        str(G) if G is not None else ""])
    wb.close()
out.close()
print("STATS:", stats)
